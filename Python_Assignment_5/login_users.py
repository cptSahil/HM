"""
Description:
This script defines a LoginManager class to manage user login operations using Selenium. 
It loads user credentials from an Excel file, performs login attempts on a web page, captures
error messages, and stores the results in another Excel sheet. The class utilizes explicit waits
for element interactions and URL changes during the login process.

Classes:
- LoginManager: Manages user login operations and result storage.

Functions/Methods:
- __init__: Initializes the LoginManager class with URL, filename, and browser manager.
- login_with_users: Performs user login attempts and stores results in an Excel file.

Dependencies:
- pandas: For data manipulation and Excel handling.
- openpyxl: For working with Excel files.
- selenium.webdriver.common.by.By: For locating elements on the web page.
- selenium.webdriver.common.keys.Keys: For keyboard actions.
- selenium.webdriver.support.ui.WebDriverWait: For explicit waits.
- selenium.webdriver.support.expected_conditions: For defining expected conditions.
- selenium.common.exceptions.TimeoutException: For handling timeout exceptions.
"""

import pandas as pd
import openpyxl
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException

class LoginManager:                               #pylint: disable=R0903
    """
    A class to manage user login operations using Selenium.

    Attributes:
    - url (str): The URL of the login page.
    - filename (str): The name of the Excel file to store login results.
    - browser_manager: An instance of BrowserManager to control the web browser.
    """

    def __init__(self, url, filename, browser_manager):
        """
        Initialize the LoginManager class.

        Parameters:
        - url (str): The URL of the login page.
        - filename (str): The name of the Excel file.
        - browser_manager: An instance of BrowserManager.
        """
        self.url = url
        self.filename = filename
        self.browser_manager = browser_manager

    def login_with_users(self):
        """
        Perform user login operations and store results in an Excel file.
        """
        self.browser_manager.setup_browser(self.url)
        try:
            workbook = openpyxl.load_workbook(self.filename)
            login_sheet = workbook.create_sheet("Login")
            login_sheet.append(["User ID", "Login Message"])
            users = workbook["User_credentials"]

            data = []
            for row in users.iter_rows(min_row=2, values_only=True):
                user_id, username, password = row
                self.browser_manager.driver.get(self.url)  # Navigate to the login page

                username_field = self.browser_manager.driver.find_element(By.ID, "user-name")
                password_field = self.browser_manager.driver.find_element(By.ID, "password")

                username_field.clear()
                password_field.clear()

                username_field.send_keys(username)
                password_field.send_keys(password)
                # time.sleep(5)
                password_field.send_keys(Keys.RETURN)
                # time.sleep(5)

                initial_url = self.browser_manager.driver.current_url

                if self.browser_manager.driver.current_url == initial_url:
                    pass
                else:
                    self.browser_manager.driver.execute_script("window.history.go(-1)")
                time.sleep(5)

                error_message = ""
                try:
                    if self.browser_manager.driver.find_elements(By.XPATH, "//h3"):
                        error_message = self.browser_manager.driver.find_element(By.XPATH, "//h3").text
                except():
                    # error_message = "No error message found"
                    pass

                data.append([user_id, error_message])
                # self.browser_manager.driver.execute_script("window.history.go(-1)")


            df = pd.DataFrame(data, columns=["User ID", "Login Message"])
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name='Login', index=False)

        finally:
            self.browser_manager.driver.quit()
