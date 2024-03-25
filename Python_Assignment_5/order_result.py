"""
Description:
This script defines a OrderPlacer class to manage user login and order placement.

Classes:
- OrderPlacer: Manages user login and place order according to data.

Functions/Methods:
- __init__: Initializes the LoginManager class with URL, filename, and browser manager.
- place_orders: Perform user login, place order, check criteria for success/failure and store data
in excel sheet.
"""
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from login import Login
from checkout_information import CheckoutInfo

class OrderPlacer:                      #pylint: disable=R0903
    """
    A class to place orders on a website using Selenium WebDriver.

    Attributes:
        filename (str): The name of the Excel file containing order details.
        url (str): The URL of the website to place orders on.
        browser_manager: An instance of the browser manager to handle WebDriver operations.
    """
    def __init__(self,filename,url,browser_manager):
        self.driver = webdriver.Chrome()
        self.url = url
        self.filename = filename
        self.browser_manager = browser_manager

    def place_orders(self):         #pylint: disable=R0914
        """
        Method to place orders based on data from an Excel file.

        This method iterates through rows in the Excel file, logs in, adds products to cart,
        fills in checkout information, and updates order status in the Excel file.
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            order_sheet = workbook["Order_details"]
            order_sheet["E1"] = "Order Status"

            row_index = 2  # Start from the second row where data begins

            for row in order_sheet.iter_rows(min_row=row_index, values_only=True):
                username, product, quantity, price, *_ = row
                print(username,product,quantity,price)

                self.browser_manager.setup_browser(self.url)
                self.browser_manager.driver.implicitly_wait(3)

                login_manager = Login(self.browser_manager,username,"secret_sauce")
                login_manager.login()

                self.browser_manager.driver.implicitly_wait(3)

                product_element = self.browser_manager.driver.find_element(By.XPATH, f"//*[text()='{product}']/../../..") #pylint: disable=C0301

                add_to_cart_button = product_element.find_element(By.CLASS_NAME, "btn_inventory")
                add_to_cart_button.click()
                self.browser_manager.driver.implicitly_wait(5)

                cart_icon = WebDriverWait(self.browser_manager.driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "shopping_cart_container"))) #pylint: disable=C0301
                cart_icon.click()
                self.browser_manager.driver.implicitly_wait(5)
                product_in_cart = self.browser_manager.driver.find_element(By.XPATH,"//div[@class='inventory_item_name']").text     #pylint: disable=C0301
                total_cost_in_cart = float(self.browser_manager.driver.find_element(By.CLASS_NAME,"inventory_item_price").text)     #pylint: disable=C0301
                # price = float(price)

                cart_quantity = self.browser_manager.driver.find_element(By.CLASS_NAME, "cart_quantity")    #pylint: disable=C0301
                cart_count = int(cart_quantity.text)
                print("Cart count:", cart_count)

                checkout_button = self.browser_manager.driver.find_element(By.CLASS_NAME, "checkout_button")    #pylint: disable=C0301
                checkout_button.click()

                checkout_manager = CheckoutInfo(self.browser_manager)
                checkout_manager.personal_info()

                continue_button = self.browser_manager.driver.find_element(By.CLASS_NAME, "cart_button")    #pylint: disable=C0301
                continue_button.click()

                finish_button = self.browser_manager.driver.find_element(By.CLASS_NAME, "cart_button")  #pylint: disable=C0301
                finish_button.click()

                cart_icon2 = WebDriverWait(self.browser_manager.driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "shopping_cart_container"))) #pylint: disable=C0301
                cart_icon2.click()
                cart_items = self.browser_manager.driver.find_elements(By.CLASS_NAME, "inventory_item_name")        #pylint: disable=C0301
                item_still_in_cart = any(product in item.text for item in cart_items)

                print(product_in_cart)
                if cart_count == int(quantity) and product_in_cart == product and total_cost_in_cart == float(price):  #pylint: disable=C0301
                    if item_still_in_cart:
                        order_sheet.cell(row=row_index, column=5, value="Failure")
                        print("Order status updated to Failure")
                    else:
                        order_sheet.cell(row=row_index, column=5, value="Success")
                        print("Order status updated to Success")

                workbook.save(self.filename)
                self.browser_manager.driver.implicitly_wait(5)
                row_index += 1
            print("All orders placed successfully!")

        finally:
            self.driver.quit()
