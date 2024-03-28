"""
Description:
This script defines a LoginManager class to manage user login operations using Selenium. 
It loads user credentials from an Excel file, performs login attempts on a web page, captures
error messages, and stores the results in another Excel sheet. The class utilizes explicit waits
for element interactions and URL changes during the login process.

Classes:
- LoginManager: Manages user login operations and result storage.

Functions/Methods:
- __init__: Initializes the LoginManager class with URL, filename, and browser manager.
- login_with_users: Performs user login attempts and stores results in an Excel file.

Dependencies:
- pandas: For data manipulation and Excel handling.
- openpyxl: For working with Excel files.
- selenium.webdriver.common.by.By: For locating elements on the web page.
- selenium.webdriver.common.keys.Keys: For keyboard actions.
- selenium.webdriver.support.ui.WebDriverWait: For explicit waits.
- selenium.webdriver.support.expected_conditions: For defining expected conditions.
- selenium.common.exceptions.TimeoutException: For handling timeout exceptions.
"""

import pandas as pd
import openpyxl
from selenium.webdriver.common.by import By
from tasks.login import Login

class DataManager:                               #pylint: disable=R0903
    """
    A class to manage user login operations using Selenium.

    Attributes:
    - url (str): The URL of the login page.
    - filename (str): The name of the Excel file to store login results.
    - browser_manager: An instance of BrowserManager to control the web browser.
    """
    error_msg_locator = "//div[@class='login-box']/form/h3"

    def __init__(self, url, filename, browser_manager):
        """
        Initialize the LoginManager class.

        Parameters:
        - url (str): The URL of the login page.
        - filename (str): The name of the Excel file.
        - browser_manager: An instance of BrowserManager.
        """
        self.url = url
        self.filename = filename
        self.browser_manager = browser_manager

    def login_with_users(self):
        """
        Perform user login operations and store results in an Excel file.
        """
        self.browser_manager.setup_browser(self.url)
        try:
            workbook = openpyxl.load_workbook(self.filename)
            login_sheet = workbook.create_sheet("Login")
            login_sheet.append(["User ID", "Login Message"])
            users = workbook["User_credentials"]

            data = []
            for row in users.iter_rows(min_row=2, values_only=True):
                user_id, username, password = row
                self.browser_manager.driver.get(self.url)

                login_manager = Login(self.browser_manager,username,password)
                login_manager.login()

                initial_url = self.browser_manager.driver.current_url

                if self.browser_manager.driver.current_url == initial_url:
                    pass
                else:
                    self.browser_manager.driver.execute_script("window.history.go(-1)")
                self.browser_manager.driver.implicitly_wait(5)

                error_message = "Login Successfully!!"
                try:
                    if self.browser_manager.driver.find_elements(By.XPATH, self.error_msg_locator):
                        error_message = self.browser_manager.driver.find_element(By.XPATH, self.error_msg_locator).text #pylint: disable=C0301
                except():
                    pass

                data.append([user_id, error_message])

            df = pd.DataFrame(data, columns=["User ID", "Login Message"])
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name='Login', index=False)

        finally:
            self.browser_manager.driver.quit()
