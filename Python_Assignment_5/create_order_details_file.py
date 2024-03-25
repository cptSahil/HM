"""
Description:

Classes:
- CreateOrderDetailsFile: Class to create the order details file.

Functions/Methods:
- __init__: Initializes the CreateOrderDetailsFile class with filename.
- save_order_details: Save order details to an Excel sheet. If the file 
  exists, it appends a new sheet 'Order_details' with columns for User,
  Product, Quantity, and Total Price.

Dependencies:
- pandas: For data manipulation and Excel handling.
- openpyxl: For working with Excel files.
"""
import pandas as pd
import openpyxl

class CreateOrderDetailsFile:                   #pylint: disable=R0903
    """
    A class to manage the creation and saving of order details in an Excel file.

    Attributes:
    - filename (str): The name of the Excel file to store order details.
    """

    def __init__(self, filename):
        """
        Initialize the CreateOrderDetailsFile class with the provided filename.

        Parameters:
        - filename (str): The name of the Excel file.
        """
        self.filename = filename

    def save_order_details(self):
        """
        Save order details to an Excel file. If the file exists, it appends a new sheet 
        'Order_details' with columns for User, Product, Quantity, and Total Price. If the 
        file doesn't exist, it creates a new Excel file with the specified columns.
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            if "Order_details" not in workbook.sheetnames:
                order_sheet = workbook.create_sheet("Order_details")
                order_sheet.append(["User", "Product", "Quantity", "Total Price"])
            workbook.save(self.filename)
        except FileNotFoundError:
            df = pd.DataFrame(columns=["User", "Product", "Quantity", "Total Price"])
            with pd.ExcelWriter(self.filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Order_details", index=False)

    def add_order_data(self, order_data_list):
        """
        Add order data to the 'Order_details' sheet in the Excel file.

        Parameters:
        - order_data (list): A list containing order details in the format 
        [User, Product, Quantity, Total Price].
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            order_sheet = workbook["Order_details"]
            for order_data in order_data_list:
                order_sheet.append(order_data)
            workbook.save(self.filename)
        except FileNotFoundError:
            print("Excel file not found. Please save order details first using save_order_details method.")
